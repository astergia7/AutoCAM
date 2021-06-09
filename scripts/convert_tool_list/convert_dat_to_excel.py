from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import openpyxl
import pandas as pd
import os
import sys

# This script converts .dat file to .xlsx format
main_data =[]
program_folder = os.path.abspath('')
Output_Table = program_folder + '\spreadsheets\\Converted_Tool_library.xlsx'
try:
    tool_Database = sys.argv[1]
    print("\n-------- dat. File from GUI Input --------\n")
except:
    tool_Database = program_folder + '\\resources\custom_tool_database\metric\\tool_database.dat'
    print("\n-------- dat. File from Script --------\n")
print('Input File: '+str(tool_Database)+'\n')

def convert_tools():
    
    with open(tool_Database, 'r') as file:
        text = file.read().split('\n')
    
    for x in range(len(text)):  
        if '#CLASS' in text[x]:
            format_string = text[x+1] # read row with tool format data
            format_string = format_string[6:] # remove "Format " part
            format_string = format_string.split()  
            i = 2 # to jump to row where tool object may be positioned
            while True:  
                if x+i < len(text): # Check if counter exceeds tool database length 
                    if 'DATA |' in text[x+i]:
                        tool_linedata ={} 
                        found_data = text[x+i]
                        found_data = found_data[6:] # remove "DATA | " part
                        tool_line= found_data.split('|')
                        tool_linedata['Usage'] = '1' # 1 - tool will be enabled by default, 0 - disabled
                        tool_linedata['Name'] = ''
                        tool_linedata['Type'] = ''
                        tool_linedata['Speed'] = ''
                        tool_linedata['Feed'] = ''
                        tool_linedata['Depth'] = ''
                        for y in range(len(format_string)):
                            tool_linedata[format_string[y]] = tool_line[y]
                        main_data.append(tool_linedata)
                        i += 1
                    elif '#END_DATA' in text[x+i]:
                        break
                    else:
                        i += 1
                else:
                    break

    for i in main_data: # remove spaces in data
        for j in i.keys():
            i[j] = i[j].replace(' ','')

    import_Table = pd.DataFrame(main_data)
    writer = pd.ExcelWriter(Output_Table, engine='xlsxwriter') 
    import_Table.to_excel(writer, sheet_name='Tools List', index=False)
    worksheet = writer.sheets['Tools List']
    workbook = writer.book
    header_cell_format = workbook.add_format()
    header_cell_format.set_align('center')
    header_cell_format.set_align('vcenter')
    col_names = [{'header': col_name} for col_name in import_Table.columns]
    worksheet.add_table(0, 0, import_Table.shape[0], import_Table.shape[1]-1, {
    'columns': col_names,
    # 'style' = option Format as table value and is case sensitive 
    # (look at the exact name into Excel)
    'style': 'Table Style Light 9'})
    for i, col in enumerate(col_names):
    # apply header_cell_format to cell on [row:0, column:i] and write text value from col_names in
        worksheet.write(0, i, col['header'], header_cell_format)
    writer.save()
    
    wb = openpyxl.load_workbook(Output_Table)
    ws = wb.active
    
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True
    wb.save(Output_Table)
    print('Output path: '+ str(Output_Table))
    print("\n-------- .xlsx File Created --------\n")

convert_tools()