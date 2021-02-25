from ncdata_builder import NxNCDataBuilder
from pathlib import Path
import os
import sys
import json
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter

if __name__ == "__main__":
    project_path = sys.argv[1]
    nx_file_name = sys.argv[2]
    nx_path = sys.argv[3]

    #project_path = 'C:/Users/vyacheslav/Desktop/FBM/feature_recognition_machining-master'
    #nx_file_name = 'Another_Assembly.prt'
    #nx_path = 'C:/Program Files/Siemens/NX1899/'
    
    nsdb = NxNCDataBuilder()
    nsdb.run(project_path,nx_file_name,nx_path)
    
    fh=open(project_path+'/output/output_data.txt')
    for line in fh:
        mypos=[pos for pos, char in enumerate(line) if char == '"']
    # print(mypos)

    c=0
    wc=0
    rownum=2 # for results rownum stars from second excel row which is 1 in python
    colnum=0
    rowinit=15
    with xlsxwriter.Workbook(project_path+'/output/nc_times.xlsx') as Workbook:
        worksheet = Workbook.add_worksheet()
        cell_format = Workbook.add_format({'bold':1,'align': 'center', 'valign': 'vcenter', 'border': 1,'font_size':11}) # Format
        worksheet.merge_range('A1:G1', str(nx_file_name)+' Machining Data', cell_format)
        worksheet.set_column('A:G', 22)
        while c<len(mypos):
            stpnt=mypos[c]+1
            c=c+1
            enpnt=mypos[c]
            c=c+1
            # print(line[stpnt:enpnt])
            wc=wc+1
            if wc == rowinit:
                rowinit = rowinit+14
                rownum = rownum +1
                colnum = 0
            
            if wc%2 == 0:
                    # list2='line[stpnt:enpnt]'
                    worksheet.write(rownum,colnum,line[stpnt:enpnt])
                    colnum=colnum+1
            else:
                    # list1='line[stpnt:enpnt]'
                    if wc>=15:
                        continue
                    else:
                        # as these are the title the rownum is equal to 0, which is first row in excel
                        worksheet.write(1,colnum,line[stpnt:enpnt],cell_format)
    
    wb = openpyxl.load_workbook(project_path+'/output/nc_times.xlsx')
    ws = wb.active
    kk = len(ws['G'])

    # Border Styles
    thin_border = openpyxl.styles.Border(left= openpyxl.styles.Side(border_style='thin',color='000000'),
                    right=openpyxl.styles.Side(border_style='thin',color='000000'),
                    top=openpyxl.styles.Side(border_style='thin',color='000000'),
                    bottom=openpyxl.styles.Side(border_style='thin',color='000000'),
                    )

    for x in range (3, kk+1):

        # "Operation name" style
        ws['A'+str(x)].border = thin_border

        # "Path time" style and recalculation:
        cell_value =  ws['G'+str(x)].value
        cell_value = float(cell_value)/(24*60)
        ws['G'+str(x)] = cell_value
        #ws['G'+str(x)] = '='+str(cell_value)+'/(24*60)'
        ws['G'+str(x)].number_format = 'mm:ss'
        ws['G'+str(x)].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws['G'+str(x)].border = thin_border

        # Tool style:
        ws['B'+str(x)].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws['B'+str(x)].border = thin_border
        
        # Stock value style:
        cell_value =  ws['C'+str(x)].value
        ws['C'+str(x)] = float(cell_value)
        ws['C'+str(x)].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws['C'+str(x)].number_format = 'General'
        ws['C'+str(x)].border = thin_border
        
        # Spindle RMP style:
        cell_value =  ws['D'+str(x)].value
        ws['D'+str(x)] = float(cell_value)
        ws['D'+str(x)].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws['D'+str(x)].number_format = '0'
        ws['D'+str(x)].border = thin_border

        # Feedrate value:
        cell_value =  ws['E'+str(x)].value
        ws['E'+str(x)] = float(cell_value)
        ws['E'+str(x)].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws['E'+str(x)].number_format = 'General'
        ws['E'+str(x)].border = thin_border
        
        # Feedrate type style:
        ws['F'+str(x)].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws['F'+str(x)].border = thin_border
    
    # Total time string:
    ws['G'+str(kk+1)] = '=SUM(G3:G'+str(kk)+')'
    ws['G'+str(kk+1)].number_format = 'mm:ss'
    ws['G'+str(kk+1)].alignment = openpyxl.styles.Alignment(horizontal='center')
    ws['G'+str(kk+1)].border = thin_border

    ws['F'+str(kk+1)] = 'Total:'
    ws['F'+str(kk+1)].font = openpyxl.styles.Font(bold=True)
    ws['F'+str(kk+1)].alignment = openpyxl.styles.Alignment(horizontal='right')
    ws['F'+str(kk+1)].border = thin_border

    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    # Delete Feedrate type row
    #ws.delete_cols(5)

    # Save file
    wb.save(project_path+'/output/nc_times.xlsx')